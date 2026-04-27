from __future__ import annotations
from _runner import run_main

print("SCRIPT STARTING", flush=True)
import argparse
import csv
import json
import logging
import sys
import traceback
from collections.abc import Iterable
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.config import settings
from app.infrastructure.db.connection import get_database_engine, get_safe_database_url

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    stream=sys.stdout,
)
logger = logging.getLogger("load_resource")

# Migration은 raw/운영 테이블을 만들고, 이 스크립트는 manifest 정의대로
# resource 파일을 해당 테이블에 적재한다.


def normalize_cell(value: Any) -> str | None:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat(sep=" ", timespec="seconds")
        except TypeError:
            return value.isoformat()
    return str(value).strip()


def read_csv_rows(path: Path) -> list[list[str | None]]:
    for encoding in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="", errors="replace") as file:
                reader = csv.reader(file)
                return [[normalize_cell(cell) for cell in row] for row in reader]
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("csv", b"", 0, 1, f"Unable to decode {path}")


def iter_xlsx_rows(path: Path, sheet_name: str | None = None) -> tuple[str, list[list[str | None]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        target_sheet = sheet_name or workbook.sheetnames[0]
        worksheet = workbook[target_sheet]
        rows = [
            [normalize_cell(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)
        ]
        return target_sheet, rows
    finally:
        workbook.close()


def insert_tabular_rows(
    connection: Any,
    *,
    table: str,
    columns: list[str],
    rows: Iterable[list[str | None]],
    source_file: str,
    source_sheet: str | None,
    loaded_at: str,
) -> int:
    all_columns = columns + ["source_file", "source_sheet", "loaded_at"]
    sql = text(
        f'INSERT INTO "{table}" ({", ".join(all_columns)}) '
        f'VALUES ({", ".join(f":{column}" for column in all_columns)})'
    )

    payload_rows = []
    for row in rows:
        if not any(value not in (None, "") for value in row):
            continue
        values = list(row[: len(columns)])
        if len(values) < len(columns):
            values.extend([None] * (len(columns) - len(values)))
        payload = dict(zip(columns, values))
        payload["source_file"] = source_file
        payload["source_sheet"] = source_sheet
        payload["loaded_at"] = loaded_at
        payload_rows.append(payload)

    if payload_rows:
        connection.execute(sql, payload_rows)
    return len(payload_rows)


def load_dataset(connection: Any, dataset: dict[str, Any], run_id: int) -> None:
    backend_root = settings.backend_root
    loaded_at = datetime.now().isoformat(timespec="seconds")

    # dataset은 db/manifests/resource_load_manifest.json의 단일 항목이다.
    for relative_path in dataset["paths"]:
        source_path = (backend_root / relative_path).resolve()
        source_file = str(source_path.relative_to(settings.project_root))
        table = dataset["table"]
        logger.info("→ loading table=%s file=%s loader=%s", table, source_file, dataset["loader"])
        connection.execute(
            text(f'DELETE FROM "{table}" WHERE source_file = :source_file'),
            {"source_file": source_file},
        )

        row_count = 0
        message = "loaded"
        status = "success"
        source_sheet = None

        try:
            if dataset["loader"] == "csv":
                rows = read_csv_rows(source_path)
                source_sheet = "csv"
                row_count = insert_tabular_rows(
                    connection,
                    table=table,
                    columns=dataset["columns"],
                    rows=rows[1:],
                    source_file=source_file,
                    source_sheet=source_sheet,
                    loaded_at=loaded_at,
                )
            elif dataset["loader"] == "xlsx":
                source_sheet, rows = iter_xlsx_rows(
                    source_path, sheet_name=dataset.get("sheet") or dataset.get("sheet_name")
                )
                row_count = insert_tabular_rows(
                    connection,
                    table=table,
                    columns=dataset["columns"],
                    rows=rows[1:],
                    source_file=source_file,
                    source_sheet=source_sheet,
                    loaded_at=loaded_at,
                )
            elif dataset["loader"] == "workbook_rows":
                workbook = load_workbook(source_path, read_only=True, data_only=True)
                try:
                    for sheet in workbook.sheetnames:
                        connection.execute(
                            text(
                                "DELETE FROM raw_workbook_rows WHERE source_file = :source_file AND sheet_name = :sheet_name"
                            ),
                            {"source_file": source_file, "sheet_name": sheet},
                        )
                        worksheet = workbook[sheet]
                        payload_rows = []
                        for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                            normalized = [normalize_cell(cell) for cell in row]
                            if not any(value not in (None, "") for value in normalized):
                                continue
                            payload_rows.append(
                                {
                                    "workbook_name": source_path.name,
                                    "sheet_name": sheet,
                                    "row_index": index,
                                    "row_values_json": json.dumps(normalized, ensure_ascii=False),
                                    "source_file": source_file,
                                    "loaded_at": loaded_at,
                                }
                            )
                            row_count += 1
                        if payload_rows:
                            connection.execute(
                                text(
                                    """
                                    INSERT INTO raw_workbook_rows(
                                        workbook_name, sheet_name, row_index, row_values_json, source_file, loaded_at
                                    ) VALUES (
                                        :workbook_name, :sheet_name, :row_index, :row_values_json, :source_file, :loaded_at
                                    )
                                    """
                                ),
                                payload_rows,
                            )
                finally:
                    workbook.close()
                source_sheet = "*"
            else:
                raise ValueError(f"Unsupported loader: {dataset['loader']}")
        except Exception as exc:
            status = "failed"
            message = str(exc)
            logger.error("✗ FAILED table=%s file=%s err=%s", table, source_file, message)
            logger.error("traceback:\n%s", traceback.format_exc())

        if status == "success":
            logger.info("✓ done table=%s rows=%d sheet=%s", table, row_count, source_sheet)

        connection.execute(
            text(
                """
                INSERT INTO ingestion_files(
                    run_id, table_name, source_file, source_sheet, row_count, loaded_at, status, message
                ) VALUES (
                    :run_id, :table_name, :source_file, :source_sheet, :row_count, :loaded_at, :status, :message
                )
                """
            ),
            {
                "run_id": run_id,
                "table_name": table,
                "source_file": source_file,
                "source_sheet": source_sheet,
                "row_count": row_count,
                "loaded_at": loaded_at,
                "status": status,
                "message": message,
            },
        )

        if status == "failed":
            raise RuntimeError(f"dataset load failed: table={table} file={source_file}: {message}")


def populate_store_clusters(connection: Any) -> int:
    column_rows = connection.execute(
        text(
            """
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND table_name = 'store_clusters'
            """
        )
    ).mappings().all()
    columns = {str(row["column_name"]).lower(): str(row["data_type"]).lower() for row in column_rows}
    required_columns = {"masked_stor_cd", "cluster_id", "cluster_label", "updated_at"}
    if not required_columns.issubset(columns):
        print(
            "Skipping store_clusters population: incompatible schema "
            f"(required={sorted(required_columns)}, actual={sorted(columns.keys())})"
        )
        return 0

    cluster_id_type = columns.get("cluster_id", "")
    if cluster_id_type not in {"text", "character varying", "character"}:
        print(
            "Skipping store_clusters population: cluster_id column type is not text "
            f"(actual={cluster_id_type})"
        )
        return 0

    insert_columns = ["masked_stor_cd"]
    select_columns = ["masked_stor_cd"]
    if "sido" in columns:
        insert_columns.append("sido")
        select_columns.append("NULLIF(TRIM(COALESCE(sido, '')), '') AS sido")
    if "store_type" in columns:
        insert_columns.append("store_type")
        select_columns.append("NULLIF(TRIM(COALESCE(store_type, '')), '') AS store_type")
    insert_columns.extend(["cluster_id", "cluster_label", "updated_at"])
    select_columns.extend(
        [
            "CONCAT("
            "COALESCE(NULLIF(TRIM(COALESCE(sido, '')), ''), 'UNKNOWN'),"
            "'|',"
            "COALESCE(NULLIF(TRIM(COALESCE(store_type, '')), ''), 'UNKNOWN')"
            ") AS cluster_id",
            "CONCAT("
            "COALESCE(NULLIF(TRIM(COALESCE(sido, '')), ''), 'UNKNOWN'),"
            "' / ',"
            "COALESCE(NULLIF(TRIM(COALESCE(store_type, '')), ''), 'UNKNOWN')"
            ") AS cluster_label",
            ":loaded_at AS updated_at",
        ]
    )

    connection.execute(text("DELETE FROM store_clusters"))
    result = connection.execute(
        text(
            f"""
            INSERT INTO store_clusters (
                {", ".join(insert_columns)}
            )
            SELECT
                {", ".join(select_columns)}
            FROM raw_store_master
            WHERE NULLIF(TRIM(COALESCE(masked_stor_cd, '')), '') IS NOT NULL
            """
        ),
        {"loaded_at": datetime.now()},
    )
    return result.rowcount or 0


def get_fifo_store_codes(engine: Any) -> list[str]:
    """FIFO 적재 대상 점포코드 목록 조회"""
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT DISTINCT masked_stor_cd FROM raw_production_extract
                WHERE NULLIF(TRIM(COALESCE(masked_stor_cd, '')), '') IS NOT NULL
                UNION
                SELECT DISTINCT masked_stor_cd FROM raw_order_extract
                WHERE NULLIF(TRIM(COALESCE(masked_stor_cd, '')), '') IS NOT NULL
                ORDER BY 1
                """
            )
        )
        return [row[0] for row in result]


def populate_fifo_lots_for_store(connection: Any, store_cd: str) -> tuple[int, int]:
    """단일 점포의 FIFO Lot 적재

    반환값: (production_rows, delivery_rows)
    """
    # 해당 점포 기존 데이터 초기화
    connection.execute(
        text("DELETE FROM inventory_fifo_lots WHERE masked_stor_cd = :s"),
        {"s": store_cd},
    )

    # 1. 생산 Lot 적재 — 유통기한 기본 1일(당일 소진 원칙)
    prod_result = connection.execute(
        text(
            """
            INSERT INTO inventory_fifo_lots
                (masked_stor_cd, item_cd, item_nm, lot_type,
                 lot_date, expiry_date, shelf_life_days,
                 initial_qty, unit_cost)
            SELECT
                p.masked_stor_cd,
                p.item_cd,
                p.item_nm,
                'production',
                TO_DATE(p.prod_dt, 'YYYYMMDD'),
                TO_DATE(p.prod_dt, 'YYYYMMDD')
                    + COALESCE(NULLIF(TRIM(s.shelf_life_days), '')::INT, 1),
                COALESCE(NULLIF(TRIM(s.shelf_life_days), '')::INT, 1),
                COALESCE(NULLIF(TRIM(p.prod_qty),   '')::NUMERIC, 0)
                + COALESCE(NULLIF(TRIM(p.prod_qty_2), '')::NUMERIC, 0)
                + COALESCE(NULLIF(TRIM(p.prod_qty_3), '')::NUMERIC, 0)
                + COALESCE(NULLIF(TRIM(p.reprod_qty), '')::NUMERIC, 0),
                COALESCE(NULLIF(TRIM(p.item_cost), '')::NUMERIC, 0)
            FROM raw_production_extract p
            LEFT JOIN raw_product_shelf_life s ON p.item_nm = s.item_nm
            WHERE p.masked_stor_cd = :s
              AND (
                COALESCE(NULLIF(TRIM(p.prod_qty),   '')::NUMERIC, 0)
                + COALESCE(NULLIF(TRIM(p.prod_qty_2), '')::NUMERIC, 0)
                + COALESCE(NULLIF(TRIM(p.prod_qty_3), '')::NUMERIC, 0)
                + COALESCE(NULLIF(TRIM(p.reprod_qty), '')::NUMERIC, 0)
              ) > 0
            """
        ),
        {"s": store_cd},
    )

    # 2. 납품 Lot 적재 — 유통기한 기본 90일(원재료 기준)
    deliv_result = connection.execute(
        text(
            """
            INSERT INTO inventory_fifo_lots
                (masked_stor_cd, item_cd, item_nm, lot_type,
                 lot_date, expiry_date, shelf_life_days,
                 initial_qty, unit_cost)
            SELECT
                o.masked_stor_cd,
                o.item_cd,
                o.item_nm,
                'delivery',
                TO_DATE(o.dlv_dt, 'YYYYMMDD'),
                TO_DATE(o.dlv_dt, 'YYYYMMDD')
                    + COALESCE(NULLIF(TRIM(s.shelf_life_days), '')::INT, 90),
                COALESCE(NULLIF(TRIM(s.shelf_life_days), '')::INT, 90),
                COALESCE(NULLIF(TRIM(o.confrm_qty), '')::NUMERIC, 0),
                COALESCE(NULLIF(TRIM(o.confrm_prc), '')::NUMERIC, 0)
            FROM raw_order_extract o
            LEFT JOIN raw_product_shelf_life s ON o.item_nm = s.item_nm
            WHERE o.masked_stor_cd = :s
              AND COALESCE(NULLIF(TRIM(o.confrm_qty), '')::NUMERIC, 0) > 0
            """
        ),
        {"s": store_cd},
    )

    # 3. FIFO 소진 — 해당 점포 판매(core_daily_item_sales) 기준 production Lot 차감
    # store_cd를 DO 블록 안에 리터럴로 삽입 (DB에서 가져온 값이므로 안전)
    safe_store_cd = store_cd.replace("'", "''")
    connection.execute(
        text(
            f"""
            DO $$
            DECLARE
                r         RECORD;
                lot       RECORD;
                remaining NUMERIC;
                deduct    NUMERIC;
            BEGIN
                FOR r IN
                    SELECT sale_dt, item_nm, sale_qty
                    FROM   core_daily_item_sales
                    WHERE  sale_qty > 0
                      AND  masked_stor_cd = '{safe_store_cd}'
                    ORDER  BY item_nm, sale_dt
                LOOP
                    remaining := r.sale_qty;

                    FOR lot IN
                        SELECT id, initial_qty, consumed_qty
                        FROM   inventory_fifo_lots
                        WHERE  masked_stor_cd = '{safe_store_cd}'
                          AND  item_nm        = r.item_nm
                          AND  lot_type       = 'production'
                          AND  lot_date       <= TO_DATE(r.sale_dt, 'YYYYMMDD')
                          AND  status         = 'active'
                        ORDER  BY lot_date ASC
                    LOOP
                        EXIT WHEN remaining <= 0;

                        deduct := LEAST(remaining, lot.initial_qty - lot.consumed_qty);
                        CONTINUE WHEN deduct <= 0;

                        UPDATE inventory_fifo_lots
                        SET    consumed_qty = consumed_qty + deduct,
                               status       = CASE
                                                WHEN consumed_qty + deduct >= initial_qty
                                                THEN 'sold_out'
                                                ELSE 'active'
                                              END,
                               updated_at   = NOW()
                        WHERE  id = lot.id;

                        remaining := remaining - deduct;
                    END LOOP;
                END LOOP;
            END;
            $$
            """
        )
    )

    # 4. 유통기한 초과 Lot 확정 — 잔여 수량을 wasted_qty에 기록
    connection.execute(
        text(
            """
            UPDATE inventory_fifo_lots
            SET    wasted_qty  = initial_qty - consumed_qty,
                   status      = 'expired',
                   updated_at  = NOW()
            WHERE  masked_stor_cd = :s
              AND  status         = 'active'
              AND  expiry_date    IS NOT NULL
              AND  expiry_date    < CURRENT_DATE
              AND  (initial_qty - consumed_qty) > 0
            """
        ),
        {"s": store_cd},
    )

    return (prod_result.rowcount or 0, deliv_result.rowcount or 0)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--start-from",
        type=int,
        default=0,
        help="manifest dataset index(0-base)부터 시작 (이전까지는 건너뜀, 디버깅용)",
    )
    args = parser.parse_args()

    logger.info("Starting load_resource_to_db (start_from=%d)", args.start_from)
    manifest = json.loads(settings.manifest_path.read_text(encoding="utf-8"))
    logger.info("Manifest loaded from %s (datasets=%d)", settings.manifest_path, len(manifest["datasets"]))
    engine = get_database_engine()
    logger.info("Engine created: %s", engine)
    if engine is None:
        raise RuntimeError(
            "PostgreSQL driver is not installed. Install psycopg before loading resource data."
        )

    logger.info("==== Phase 1: manifest dataset 적재 시작 ====")
    with engine.begin() as connection:
        logger.info("Transaction started. Inserting ingestion_run...")
        run_id = connection.execute(
            text(
                """
                INSERT INTO ingestion_runs(started_at, status, message)
                VALUES (:started_at, :status, :message)
                RETURNING run_id
                """
            ),
            {
                "started_at": datetime.now(),
                "status": "running",
                "message": "resource ingestion started",
            },
        ).scalar_one()

        try:
            for idx, dataset in enumerate(manifest["datasets"]):
                if idx < args.start_from:
                    logger.info("[%d/%d] SKIP table=%s (start_from=%d)", idx, len(manifest["datasets"]), dataset["table"], args.start_from)
                    continue
                logger.info("[%d/%d] dataset start table=%s", idx, len(manifest["datasets"]), dataset["table"])
                load_dataset(connection, dataset, run_id)
            logger.info("==== Phase 1.5: store_clusters 파생 ====")
            cluster_row_count = populate_store_clusters(connection)
            logger.info("store_clusters populated rows=%d", cluster_row_count)
            connection.execute(
                text(
                    """
                    INSERT INTO ingestion_files(
                        run_id, table_name, source_file, source_sheet, row_count, loaded_at, status, message
                    ) VALUES (
                        :run_id, :table_name, :source_file, :source_sheet, :row_count, :loaded_at, :status, :message
                    )
                    """
                ),
                {
                    "run_id": run_id,
                    "table_name": "store_clusters",
                    "source_file": "derived:raw_store_master",
                    "source_sheet": None,
                    "row_count": cluster_row_count,
                    "loaded_at": datetime.now().isoformat(timespec="seconds"),
                    "status": "success",
                    "message": "store clusters populated from raw_store_master",
                },
            )
            connection.execute(
                text(
                    """
                    UPDATE ingestion_runs
                    SET completed_at = :completed_at, status = :status, message = :message
                    WHERE run_id = :run_id
                    """
                ),
                {
                    "completed_at": datetime.now(),
                    "status": "success",
                    "message": "resource ingestion completed",
                    "run_id": run_id,
                },
            )
        except Exception as exc:
            logger.error("Phase 1 FAILED run_id=%s err=%s", run_id, exc)
            logger.error("traceback:\n%s", traceback.format_exc())
            connection.execute(
                text(
                    """
                    UPDATE ingestion_runs
                    SET completed_at = :completed_at, status = :status, message = :message
                    WHERE run_id = :run_id
                    """
                ),
                {
                    "completed_at": datetime.now(),
                    "status": "failed",
                    "message": str(exc),
                    "run_id": run_id,
                },
            )
            raise

    logger.info("Loaded resource data into %s", get_safe_database_url())

    # Phase 2: FIFO Lot 적재 — 점포별 별도 트랜잭션
    logger.info("==== Phase 2: FIFO lot population per store ====")
    store_codes = get_fifo_store_codes(engine)
    total_prod = total_deliv = 0
    failed_stores: list[tuple[str, str]] = []
    for idx, store_cd in enumerate(store_codes, 1):
        try:
            with engine.begin() as conn:
                prod_cnt, deliv_cnt = populate_fifo_lots_for_store(conn, store_cd)
            total_prod += prod_cnt
            total_deliv += deliv_cnt
            logger.info("  [%d/%d] %s: prod=%d, deliv=%d", idx, len(store_codes), store_cd, prod_cnt, deliv_cnt)
        except Exception as exc:
            failed_stores.append((store_cd, str(exc)))
            logger.error("  [%d/%d] %s FAILED: %s", idx, len(store_codes), store_cd, exc)
            logger.error("traceback:\n%s", traceback.format_exc())

    logger.info(
        "FIFO lots complete — stores=%d, ok=%d, failed=%d, production=%d, delivery=%d",
        len(store_codes), len(store_codes) - len(failed_stores), len(failed_stores),
        total_prod, total_deliv,
    )
    if failed_stores:
        logger.error("Failed stores summary:")
        for store_cd, err in failed_stores:
            logger.error("  - %s: %s", store_cd, err)
        sys.exit(2)


if __name__ == "__main__":
    run_main(main)